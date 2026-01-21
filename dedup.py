#!/usr/bin/env python3
"""
猎聘查重数据生成工具
支持从剪贴板、文本文件或交互式输入读取候选人数据，自动生成 Excel 查重文件
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from rich.console import Console
from rich.prompt import Prompt, Confirm
from rich.table import Table
from rich.panel import Panel
import sys
import re
import shutil
from pathlib import Path
from datetime import datetime

# pyperclip 是可选依赖，仅在使用剪贴板功能时需要
try:
    import pyperclip
    PYPERCLIP_AVAILABLE = True
except ImportError:
    PYPERCLIP_AVAILABLE = False

console = Console()


def parse_candidate_line(line):
    """
    解析候选人数据行
    支持格式: 姓名\t公司\t职位\t电话\t在职时间
    或: 姓名 公司 职位 电话 在职时间 (空格/制表符分隔)
    """
    line = line.strip()
    if not line:
        return None
    
    # 尝试按制表符分隔
    parts = line.split('\t')
    if len(parts) < 3:
        # 尝试按多个空格分隔
        parts = re.split(r'\s{2,}', line)
    if len(parts) < 3:
        # 尝试按单个空格分隔
        parts = line.split()
    
    # 至少需要: 姓名、公司、职位
    if len(parts) < 3:
        return None
    
    name = parts[0].strip()
    company = parts[1].strip()
    position = parts[2].strip()
    phone = parts[3].strip() if len(parts) > 3 else ""
    work_time = parts[4].strip() if len(parts) > 4 else ""
    
    return {
        "姓名": name,
        "在职公司": company,
        "职位": position,
        "云号码": phone,
        "在职时间": work_time
    }


def read_from_clipboard():
    """从剪贴板读取数据"""
    if not PYPERCLIP_AVAILABLE:
        console.print("[red]✗ 剪贴板功能不可用: 缺少 pyperclip 库[/red]")
        console.print("[yellow]💡 安装方法: pip3 install --break-system-packages pyperclip[/yellow]")
        console.print("[yellow]💡 或者选择其他输入方式（文件/交互式）[/yellow]")
        return None
    
    try:
        content = pyperclip.paste()
        if not content.strip():
            console.print("[yellow]⚠️  剪贴板为空[/yellow]")
            return None
        return content
    except Exception as e:
        console.print(f"[red]✗ 读取剪贴板失败: {e}[/red]")
        return None


def read_from_file(filepath):
    """从文本文件读取数据"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        return content
    except FileNotFoundError:
        console.print(f"[red]✗ 文件不存在: {filepath}[/red]")
        return None
    except Exception as e:
        console.print(f"[red]✗ 读取文件失败: {e}[/red]")
        return None


def read_from_input():
    """交互式输入数据"""
    console.print("\n[cyan]请输入候选人数据 (每行一个，格式: 姓名 公司 职位 电话 在职时间)[/cyan]")
    console.print("[dim]提示: 输入空行结束输入[/dim]\n")
    
    lines = []
    while True:
        try:
            line = input()
            if not line.strip():
                break
            lines.append(line)
        except EOFError:
            break
    
    return '\n'.join(lines) if lines else None


def parse_content(content):
    """解析内容为候选人列表"""
    lines = content.strip().split('\n')
    candidates = []
    current_category = None
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # 检查是否是分类行（只有一个词，且不包含数字）
        if len(line.split()) == 1 and not any(char.isdigit() for char in line):
            current_category = line
            continue
        
        # 解析候选人数据
        candidate = parse_candidate_line(line)
        if candidate:
            # 如果有分类，添加到候选人数据中
            if current_category:
                candidate["分类"] = current_category
            candidates.append(candidate)
    
    return candidates


def create_excel(candidates, output_path, category_name=None):
    """创建 Excel 文件"""
    if not candidates:
        console.print("[red]✗ 没有有效的候选人数据[/red]")
        return False
    
    # 确保所有候选人都有分类
    for candidate in candidates:
        if "分类" not in candidate or not candidate["分类"]:
            if category_name:
                candidate["分类"] = category_name
            else:
                # 尝试从公司名推断分类
                candidate["分类"] = candidate.get("在职公司", "未分类")
    
    # 创建 DataFrame
    df = pd.DataFrame(candidates)
    
    # 添加序号
    df.insert(0, "序号", range(1, len(df) + 1))
    
    # 添加公司列（与分类相同，用于查重）
    df["公司"] = df["分类"]
    
    # 添加空列
    df["Profile"] = ""
    df["简历链接"] = ""
    df["是否合作"] = ""
    df["最后一次登录时间"] = ""
    
    # 重新排列列顺序
    df = df[["序号", "分类", "公司", "姓名", "在职公司", "职位", "云号码", "在职时间", 
             "Profile", "简历链接", "是否合作", "最后一次登录时间"]]
    
    # 创建 Excel 文件
    wb = Workbook()
    ws = wb.active
    ws.title = "查重数据"
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 表头格式
            if r_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 调整列宽
    column_widths = {
        "A": 8,   # 序号
        "B": 12,  # 分类
        "C": 12,  # 公司
        "D": 15,  # 姓名
        "E": 25,  # 在职公司
        "F": 25,  # 职位
        "G": 18,  # 云号码
        "H": 18,  # 在职时间
        "I": 40,  # Profile
        "J": 15,  # 简历链接
        "K": 12,  # 是否合作
        "L": 20,  # 最后一次登录时间
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 保存文件
    try:
        wb.save(output_path)
        return True
    except Exception as e:
        console.print(f"[red]✗ 保存文件失败: {e}[/red]")
        return False


def display_preview(candidates):
    """显示候选人预览"""
    if not candidates:
        return
    
    table = Table(title="📋 候选人数据预览", show_lines=True)
    table.add_column("序号", style="cyan", width=6)
    table.add_column("分类", style="magenta", width=12)
    table.add_column("姓名", style="green", width=15)
    table.add_column("在职公司", style="yellow", width=25)
    table.add_column("职位", style="blue", width=20)
    table.add_column("在职时间", style="white", width=15)
    
    for idx, candidate in enumerate(candidates[:10], 1):  # 只显示前10条
        table.add_row(
            str(idx),
            candidate.get("分类", ""),
            candidate.get("姓名", ""),
            candidate.get("在职公司", ""),
            candidate.get("职位", ""),
            candidate.get("在职时间", "")
        )
    
    if len(candidates) > 10:
        table.add_row("...", "...", "...", "...", "...", "...", style="dim")
    
    console.print(table)
    console.print(f"\n[cyan]总计: {len(candidates)} 条候选人数据[/cyan]\n")


def clear_output_directories():
    """清空输出目录"""
    dirs_to_clear = ['data', 'resumes', 'zips']
    console.print("\n[yellow]--- 正在清空输出目录... ---[/yellow]")
    for directory in dirs_to_clear:
        if Path(directory).exists():
            try:
                for item in Path(directory).iterdir():
                    if item.is_file() or item.is_symlink():
                        item.unlink()
                    elif item.is_dir():
                        shutil.rmtree(item)
                console.print(f"[green]--- 已清空: {directory}/ ---[/green]")
            except Exception as e:
                console.print(f"[red]--- 清空 {directory}/ 失败: {e} ---[/red]")
        else:
            console.print(f"[dim]--- 目录不存在，跳过: {directory}/ ---[/dim]")
    console.print("[green]--- 清空完成 ---[/green]\n")


def archive_output_directories():
    """归档输出目录中的旧文件"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_name = f"archive_{timestamp}"
    dirs_to_archive = ['data', 'resumes', 'zips']
    
    console.print(f"\n[yellow]--- 正在归档旧文件到 {archive_name}... ---[/yellow]")
    
    for directory in dirs_to_archive:
        dir_path = Path(directory)
        if not dir_path.exists():
            continue
        archive_path = dir_path / archive_name
        archive_path.mkdir(exist_ok=True)
        
        try:
            for item in dir_path.iterdir():
                if item.name == archive_name:
                    continue
                shutil.move(str(item), str(archive_path / item.name))
            console.print(f"[green]--- 已归档 {directory}/ 内容 ---[/green]")
        except Exception as e:
            console.print(f"[red]--- 归档 {directory}/ 失败: {e} ---[/red]")
    console.print("[green]--- 归档完成 ---[/green]\n")


def main():
    """主函数"""
    console.print(Panel.fit(
        "[bold cyan]猎聘查重数据生成工具[/bold cyan]\n"
        "[dim]支持从剪贴板、文本文件或交互式输入读取候选人数据[/dim]",
        border_style="cyan"
    ))
    
    # 文件夹清理选项
    if Confirm.ask("是否清空 data, resumes, zips 文件夹下的所有内容? (y=清空, n=归档)", default=False):
        clear_output_directories()
    else:
        archive_output_directories()
    
    # 选择数据来源
    console.print("\n[bold]请选择数据来源:[/bold]")
    console.print("  [cyan]1[/cyan] - 从剪贴板读取")
    console.print("  [cyan]2[/cyan] - 从文本文件读取")
    console.print("  [cyan]3[/cyan] - 交互式输入")
    
    choice = Prompt.ask("请选择", choices=["1", "2", "3"], default="1")
    
    content = None
    if choice == "1":
        console.print("\n[cyan]📋 正在从剪贴板读取数据...[/cyan]")
        content = read_from_clipboard()
    elif choice == "2":
        filepath = Prompt.ask("\n请输入文件路径")
        console.print(f"\n[cyan]📄 正在从文件读取数据: {filepath}[/cyan]")
        content = read_from_file(filepath)
    else:
        content = read_from_input()
    
    if not content:
        console.print("[red]✗ 没有读取到任何数据[/red]")
        sys.exit(1)
    
    # 解析数据
    console.print("\n[cyan]🔍 正在解析数据...[/cyan]")
    candidates = parse_content(content)
    
    if not candidates:
        console.print("[red]✗ 未能解析出有效的候选人数据[/red]")
        console.print("[yellow]请确保数据格式正确: 姓名 公司 职位 电话 在职时间[/yellow]")
        sys.exit(1)
    
    # 显示预览
    display_preview(candidates)
    
    # 确认是否继续
    if not Confirm.ask("是否继续生成 Excel 文件?", default=True):
        console.print("[yellow]已取消[/yellow]")
        sys.exit(0)
    
    # 询问分类名称（可选）
    category_name = Prompt.ask(
        "\n请输入分类名称 (可选，直接回车跳过)",
        default=""
    )
    
    # 询问输出文件名
    default_filename = f"{category_name}-查重数据.xlsx" if category_name else "查重数据.xlsx"
    filename = Prompt.ask(
        "请输入输出文件名",
        default=default_filename
    )
    
    # 确保文件名以 .xlsx 结尾
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    # 输出路径
    output_dir = Path("./data")
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / filename
    
    # 生成 Excel
    console.print(f"\n[cyan]📊 正在生成 Excel 文件...[/cyan]")
    if create_excel(candidates, output_path, category_name):
        console.print(f"\n[bold green]✅ Excel 文件创建成功![/bold green]")
        console.print(f"[cyan]📁 文件路径: {output_path}[/cyan]")
        console.print(f"[cyan]📊 候选人数量: {len(candidates)}[/cyan]")
        
        # 按分类统计
        categories = {}
        for candidate in candidates:
            cat = candidate.get("分类", "未分类")
            categories[cat] = categories.get(cat, 0) + 1
        
        if len(categories) > 1:
            console.print("\n[bold]分类统计:[/bold]")
            for cat, count in categories.items():
                console.print(f"  [cyan]{cat}[/cyan]: {count} 人")
    else:
        console.print("[red]✗ 生成失败[/red]")
        sys.exit(1)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n[yellow]已取消[/yellow]")
        sys.exit(0)
    except Exception as e:
        console.print(f"\n[red]✗ 发生错误: {e}[/red]")
        import traceback
        traceback.print_exc()
        sys.exit(1)
